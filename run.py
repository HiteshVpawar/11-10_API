# run.py
from flask import Flask, request, jsonify,render_template
from app.controllers.user_controller import UserController
from app.controllers.auth_controller import AuthController
from app.controllers.chat_controller import ChatController
from app.controllers.upload_Controller import UploadController
from app.views.user_view import UserView
from app.views.auth_view import AuthView
from app.views.chat_view import ChatView
from app.views.upload_view import UploadView
from flask_cors import CORS, cross_origin
import jwt
from PyPDF2 import PdfReader
from openpyxl import load_workbook

     

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your_secret_key'
CORS(app)

# Initialize controller and view
user_controller = UserController()
auth_controller = AuthController()
chat_controller = ChatController()
upload_controller=UploadController()
user_view = UserView(user_controller)
auth_view = AuthView(auth_controller)
chat_view = ChatView(chat_controller)
upload_view=UploadView(upload_controller)

# Routes

@app.route('/auth/signup', methods=['POST'])
def signup():
        data = request.get_json()
        username = data.get('username')
        email = data.get('email')
        password = data.get('password')
        confirmPassword = data.get('confirmPassword')
        return auth_view.Sign_up(username,email,password,confirmPassword)
    #     return jsonify({"message": "Signup successful"}), 200
    # except Exception as e:
    #     return jsonify({"error": str(e)}), 500

@app.route('/auth/login', methods=['POST'])
def login():
    data = request.get_json()
    username = data.get('username')
    password = data.get('password')
    return auth_view.login(username, password)

@app.route('/hello', methods=['GET'])
def hello():
    return jsonify({'message':'Hello World Using Python Rest API !'}), 200

@app.route('/users', methods=['POST'])
def create_user():
    data = request.get_json()
    username = data.get('username')
    email = data.get('email')
    password = data.get('password')
    return user_view.create_user(username, email, password)

def isTokenValid(request):
    return auth_view.verify_token(request.headers.get('Authorization'))

@app.route('/users', methods=['GET'])
def get_users():
    response = {}
    token_res = isTokenValid(request)
    if token_res['isTokenExpired'] == True:
        return {'message':'Unauthorized','data':'Access token expired'}, 401
    elif token_res['isTokenInvalid'] == True:
        return {'message':'Unauthorized','data':'Invalid access token'}, 401
    elif len(token_res['token'])>0:
        return user_view.get_users()
    else:
        return {'message':'Unable to process the request'}, 501

@app.route('/users', methods=['PUT'])
def update_user():
    user_id = request.args.get('user_id')
    data = request.get_json()
    return user_view.update_user(user_id, data.get('username'), data.get('email'))

@app.route('/users', methods=['DELETE'])
def delete_user():
    user_id = request.args.get('user_id')
    user_deleted = user_view.delete_user(user_id)
    return user_deleted

# @app.route('/chat', methods=['POST'])
# def chat():
#     return chat_view.chat(request)

# @app.route('/chat', methods=['GET'])
# def get_chats():
#     return chat_view.get_chats()


@app.route('/chat/<user_id>', methods=['POST'])
def chat(user_id):
    return chat_view.chat(user_id, request)


@app.route('/chat/<user_id>', methods=['GET'])
def get_chats(user_id):
    return chat_view.get_chats(user_id)


@app.route('/upload', methods=['POST'])
def upload_file():
    if request.method == 'POST':
        pdf_file = request.files['pdf_file']
        if pdf_file:
            pdf_text = upload_view.extract_text_from_pdf(pdf_file)
            return pdf_text
        else:
            return "No file selected"
        

@app.route('/upload', methods=['GET'])
def get_Text():
    return upload_view.get_text()
     

@app.route('/upload/image', methods=['POST'])
def upload_image():
    if request.method == 'POST':
        image_file = request.files['image_file']
        if image_file:
            image_text = upload_view.extract_text_from_image(image_file)
            
            return image_text
        else:
            return "No file selected" 
           
@app.route('/upload/to_excel', methods=['POST'])
def upload_to_excel_endpoint():
    if request.method == 'POST':
        excel_file=request.files['excel_file']
        if excel_file :
            excel_text=upload_view.upload_to_excel(excel_file)
        # data = request.get_json()  # Assuming data is sent as JSON
        # response = upload_view.upload_to_excel(data)
        return excel_text
    else :
        return "no file selected"
        
@app.route('/extract_data', methods=['POST'])
def extract_data():
    try:
        excel_file = request.files['excel_file']
        if excel_file and excel_file.filename.endswith(('.xlsx', '.xls')):
            workbook = load_workbook(excel_file, read_only=True)
            sheet = workbook.active
            data = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                data.append(row)

            return jsonify({"data": data})
        else:
            return jsonify({"error": "Invalid file format. Please upload an Excel file."})
    except Exception as e:
        return jsonify({"error": str(e)})



if __name__ == '__main__':
    app.run(debug=True)
